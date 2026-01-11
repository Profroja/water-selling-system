from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from .models import Customer, Street
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Create your views here.

@login_required(login_url='login')
def customer_list(request):
    if request.user.role not in ['manager', 'staff']:
        return redirect('login')
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'add_street':
            region = request.POST.get('region')
            district = request.POST.get('district')
            street_name = request.POST.get('street_name')
            
            if region and district and street_name:
                Street.objects.get_or_create(
                    region=region,
                    district=district,
                    street_name=street_name
                )
                messages.success(request, f'Street "{street_name}" added successfully.')
            else:
                messages.error(request, 'Please fill in all street fields.')
            return redirect('customer_list')
        
        elif action == 'delete_customer':
            customer_id = request.POST.get('customer_id')
            if customer_id:
                try:
                    customer = Customer.objects.get(id=customer_id)
                    customer_name = f"{customer.first_name} {customer.last_name}".strip()
                    customer.delete()
                    messages.success(request, f'Mteja "{customer_name}" amefutwa.')
                except Customer.DoesNotExist:
                    messages.error(request, 'Mteja hajapatikana.')
            return redirect('customer_list')
        
        else:
            first_name = request.POST.get('first_name')
            last_name = request.POST.get('last_name')
            phone_number = request.POST.get('phone_number')
            street_id = request.POST.get('street')
            has_mita = request.POST.get('has_mita', 'no')
            total_units = request.POST.get('total_units', '0')
            
            if first_name and phone_number and street_id:
                street = Street.objects.get(id=street_id)
                
                # Parse total_units - default to 0 if not provided or invalid
                try:
                    units_value = float(total_units) if has_mita == 'yes' and total_units else 0
                except (ValueError, TypeError):
                    units_value = 0
                
                Customer.objects.create(
                    first_name=first_name,
                    last_name=last_name or '',
                    phone_number=phone_number,
                    street=street,
                    total_units=units_value
                )
                messages.success(request, f'Customer "{first_name} {last_name}" added successfully.')
                return redirect('customer_list')
            else:
                messages.error(request, 'Please fill in all required fields.')
    
    customers = Customer.objects.select_related('street').all()
    streets = Street.objects.all()
    
    # Calculate customer stats
    total_customers = customers.count()
    customers_with_units = customers.filter(total_units__gt=0).count()
    customers_without_units = customers.filter(total_units=0).count()
    
    return render(request, 'customers.html', {
        'customers': customers,
        'streets': streets,
        'active_page': 'customers',
        'total_customers': total_customers,
        'customers_with_units': customers_with_units,
        'customers_without_units': customers_without_units,
    })


@login_required(login_url='login')
def street_list(request):
    if request.user.role not in ['manager', 'staff']:
        return redirect('login')
    
    if request.method == 'POST':
        region = request.POST.get('region')
        district = request.POST.get('district')
        street_name = request.POST.get('street_name')
        
        if region and district and street_name:
            Street.objects.get_or_create(
                region=region,
                district=district,
                street_name=street_name
            )
            messages.success(request, f'Street "{street_name}" added successfully.')
        else:
            messages.error(request, 'Please fill in all required fields.')
        return redirect('street_list')
    
    streets = Street.objects.all()
    return render(request, 'streets.html', {
        'streets': streets,
        'active_page': 'streets'
    })


@login_required(login_url='login')
def staff_download_customer_template(request):
    """Download Excel template for bulk customer import - Staff version"""
    if request.user.role not in ['manager', 'staff']:
        return redirect('login')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Wateja Template"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ['Jina la Kwanza', 'Jina la Mwisho', 'Namba ya Simu', 'Region', 'District', 'Mtaa (Street)', 'Mita No. (Units)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Set column widths
    column_widths = [20, 20, 18, 15, 15, 20, 18]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Add example row
    example_data = ['Juma', 'Hassan', '0712345678', 'Dar es Salaam', 'Kinondoni', 'Mwenge', '0']
    for col, value in enumerate(example_data, 1):
        cell = ws.cell(row=2, column=col, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
    
    # Add instructions row
    ws.cell(row=4, column=1, value="MAELEKEZO:")
    ws.cell(row=4, column=1).font = Font(bold=True, color="FF0000")
    ws.cell(row=5, column=1, value="1. Jina la Kwanza ni lazima (required)")
    ws.cell(row=6, column=1, value="2. Jina la Mwisho si lazima (optional)")
    ws.cell(row=7, column=1, value="3. Namba ya Simu ni lazima (required)")
    ws.cell(row=8, column=1, value="4. Region, District, na Mtaa - tumia orodha kwenye sheet 'Mitaa Iliyopo'")
    ws.cell(row=9, column=1, value="5. Mita No. weka 0 kama mteja hana mita")
    ws.cell(row=10, column=1, value="6. Futa mfano wa row 2 kabla ya kuupload")
    ws.cell(row=11, column=1, value="7. Angalia sheet 'Mitaa Iliyopo' kuona mitaa yote iliyopo")
    
    # Create second sheet with available streets
    ws_streets = wb.create_sheet(title="Mitaa Iliyopo")
    
    # Street sheet headers
    street_headers = ['S/N', 'Region', 'District', 'Mtaa (Street)']
    street_header_fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
    for col, header in enumerate(street_headers, 1):
        cell = ws_streets.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = street_header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Set street sheet column widths
    street_col_widths = [8, 20, 20, 25]
    for col, width in enumerate(street_col_widths, 1):
        ws_streets.column_dimensions[get_column_letter(col)].width = width
    
    # Add all available streets
    streets = Street.objects.all().order_by('region', 'district', 'street_name')
    for idx, street in enumerate(streets, 1):
        row_num = idx + 1
        ws_streets.cell(row=row_num, column=1, value=idx).border = thin_border
        ws_streets.cell(row=row_num, column=2, value=street.region).border = thin_border
        ws_streets.cell(row=row_num, column=3, value=street.district).border = thin_border
        ws_streets.cell(row=row_num, column=4, value=street.street_name).border = thin_border
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="wateja_template.xlsx"'
    wb.save(response)
    return response


@login_required(login_url='login')
def staff_upload_customers_excel(request):
    """Upload Excel file to bulk import customers - Staff version"""
    if request.user.role not in ['manager', 'staff']:
        return redirect('login')
    
    if request.method != 'POST':
        return redirect('customer_list')
    
    excel_file = request.FILES.get('excel_file')
    if not excel_file:
        messages.error(request, 'Tafadhali chagua faili ya Excel.')
        return redirect('customer_list')
    
    if not excel_file.name.endswith(('.xlsx', '.xls')):
        messages.error(request, 'Tafadhali upload faili ya Excel (.xlsx au .xls).')
        return redirect('customer_list')
    
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        success_count = 0
        error_count = 0
        errors = []
        
        # Skip header row, start from row 2
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Skip empty rows
            if not row or not any(row):
                continue
            
            first_name = str(row[0]).strip() if row[0] else ''
            last_name = str(row[1]).strip() if row[1] else ''
            phone_number = str(row[2]).strip() if row[2] else ''
            region = str(row[3]).strip() if row[3] else ''
            district = str(row[4]).strip() if row[4] else ''
            street_name = str(row[5]).strip() if row[5] else ''
            
            try:
                total_units = float(row[6]) if row[6] else 0
            except (ValueError, TypeError):
                total_units = 0
            
            # Validate required fields
            if not first_name:
                errors.append(f"Row {row_num}: Jina la Kwanza halijajazwa")
                error_count += 1
                continue
            
            if not phone_number:
                errors.append(f"Row {row_num}: Namba ya Simu haijajazwa")
                error_count += 1
                continue
            
            if not region or not district or not street_name:
                errors.append(f"Row {row_num}: Region, District, au Mtaa haijajazwa")
                error_count += 1
                continue
            
            # Get or create street
            street, _ = Street.objects.get_or_create(
                region=region,
                district=district,
                street_name=street_name
            )
            
            # Create customer
            Customer.objects.create(
                first_name=first_name,
                last_name=last_name,
                phone_number=phone_number,
                street=street,
                total_units=total_units
            )
            success_count += 1
        
        if success_count > 0:
            messages.success(request, f'Wateja {success_count} wameongezwa kwa mafanikio!')
        
        if error_count > 0:
            error_msg = f'Wateja {error_count} hawakuongezwa. '
            if errors[:5]:
                error_msg += 'Makosa: ' + '; '.join(errors[:5])
                if len(errors) > 5:
                    error_msg += f' ... na mengine {len(errors) - 5}'
            messages.warning(request, error_msg)
        
    except Exception as e:
        messages.error(request, f'Hitilafu wakati wa kusoma faili: {str(e)}')
    
    return redirect('customer_list')
