from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from datetime import timedelta
from decimal import Decimal
from auths.models import CustomUser
from customers.models import Customer, Street, CustomerDebt, DebtPayment
from sales_billing.models import Order, Expenditure, CustomReportEntry
from water_production.models import WaterProduction
from unit_configs.models import UnitConfig
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Create your views here.

@login_required(login_url='login')
def manager_dashboard(request):
    if request.user.role != 'manager':
        return redirect('login')
    
    # Get week offset from query params (0 = current week, 1 = last week, etc.)
    week_offset = int(request.GET.get('week', 0))
    
    # Calculate week start and end based on offset
    today = timezone.now().date()
    current_week_start = today - timedelta(days=today.weekday())  # Monday of current week
    week_start = current_week_start - timedelta(weeks=week_offset)
    week_end = week_start + timedelta(days=6)  # Sunday
    
    # Get ISO week number and year
    week_number = week_start.isocalendar()[1]
    year = week_start.isocalendar()[0]
    
    # Get orders for the selected week
    orders = Order.objects.filter(
        created_at__date__gte=week_start,
        created_at__date__lte=week_end
    )
    order_stats = orders.aggregate(
        total_units=Sum('units'),
        total_sales=Sum('total_amount')
    )
    total_units = order_stats['total_units'] or 0
    total_sales = order_stats['total_sales'] or 0
    total_orders = orders.count()
    
    # Get expenditure for the selected week
    expenditures = Expenditure.objects.filter(
        created_at__date__gte=week_start,
        created_at__date__lte=week_end
    )
    expenditure_total = expenditures.aggregate(total=Sum('amount'))
    total_expenditure = expenditure_total['total'] or 0
    
    # Calculate profit
    profit = total_sales - total_expenditure
    
    return render(request, 'manager_dashboard.html', {
        'active_page': 'dashboard',
        'week_offset': week_offset,
        'week_number': week_number,
        'year': year,
        'week_start': week_start,
        'week_end': week_end,
        'total_units': total_units,
        'total_sales': total_sales,
        'total_orders': total_orders,
        'total_expenditure': total_expenditure,
        'profit': profit,
    })


@login_required(login_url='login')
def staff_list(request):
    if request.user.role != 'manager':
        return redirect('login')
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        # Handle toggle active status
        if action == 'toggle_status':
            user_id = request.POST.get('user_id')
            try:
                user = CustomUser.objects.get(id=user_id)
                user.is_active = not user.is_active
                user.save()
                status = 'activated' if user.is_active else 'deactivated'
                messages.success(request, f'User "{user.username}" has been {status}.')
            except CustomUser.DoesNotExist:
                messages.error(request, 'User not found.')
            return redirect('staff_list')
        
        # Handle edit user
        if action == 'edit':
            user_id = request.POST.get('user_id')
            username = request.POST.get('username')
            email = request.POST.get('email')
            phone_number = request.POST.get('phone_number')
            role = request.POST.get('role')
            new_password = request.POST.get('new_password')
            
            try:
                user = CustomUser.objects.get(id=user_id)
                
                # Check if username is taken by another user
                if CustomUser.objects.filter(username=username).exclude(id=user_id).exists():
                    messages.error(request, 'Username already exists.')
                    return redirect('staff_list')
                
                # Check if email is taken by another user
                if email and CustomUser.objects.filter(email=email).exclude(id=user_id).exists():
                    messages.error(request, 'Email already exists.')
                    return redirect('staff_list')
                
                user.username = username
                user.email = email
                user.phone_number = phone_number
                user.role = role
                
                # Update password if provided
                if new_password:
                    user.set_password(new_password)
                
                user.save()
                messages.success(request, f'User "{username}" updated successfully.')
            except CustomUser.DoesNotExist:
                messages.error(request, 'User not found.')
            return redirect('staff_list')
        
        # Handle delete user
        if action == 'delete':
            user_id = request.POST.get('user_id')
            try:
                user = CustomUser.objects.get(id=user_id)
                username = user.username
                user.delete()
                messages.success(request, f'User "{username}" has been deleted.')
            except CustomUser.DoesNotExist:
                messages.error(request, 'User not found.')
            return redirect('staff_list')
        
        # Handle add new user (default action)
        username = request.POST.get('username')
        email = request.POST.get('email')
        phone_number = request.POST.get('phone_number')
        password = request.POST.get('password')
        role = request.POST.get('role')
        
        if username and password and role:
            # Check if username already exists
            if CustomUser.objects.filter(username=username).exists():
                messages.error(request, 'Username already exists.')
            elif email and CustomUser.objects.filter(email=email).exists():
                messages.error(request, 'Email already exists.')
            else:
                user = CustomUser.objects.create_user(
                    username=username,
                    email=email,
                    password=password,
                    phone_number=phone_number,
                    role=role
                )
                messages.success(request, f'Staff member "{username}" created successfully.')
                return redirect('staff_list')
    
    # Get all managers and staff
    staff_list = CustomUser.objects.filter(role__in=['manager', 'staff']).order_by('-date_joined')
    
    return render(request, 'staff.html', {'staff_list': staff_list, 'active_page': 'staff'})


@login_required(login_url='login')
def manager_customer_list(request):
    if request.user.role != 'manager':
        return redirect('login')
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'add_street':
            region = request.POST.get('region')
            district = request.POST.get('district')
            street_name = request.POST.get('street_name')
            
            if region and district and street_name:
                Street.objects.get_or_create(
                    region=region,
                    district=district,
                    street_name=street_name
                )
                messages.success(request, f'Street "{street_name}" added successfully.')
            else:
                messages.error(request, 'Please fill in all street fields.')
            return redirect('manager_customer_list')
        
        elif action == 'delete_customer':
            customer_id = request.POST.get('customer_id')
            if customer_id:
                try:
                    customer = Customer.objects.get(id=customer_id)
                    customer_name = f"{customer.first_name} {customer.last_name}".strip()
                    customer.delete()
                    messages.success(request, f'Mteja "{customer_name}" amefutwa.')
                except Customer.DoesNotExist:
                    messages.error(request, 'Mteja hajapatikana.')
            return redirect('manager_customer_list')
        
        else:
            first_name = request.POST.get('first_name')
            last_name = request.POST.get('last_name')
            phone_number = request.POST.get('phone_number')
            street_id = request.POST.get('street')
            has_mita = request.POST.get('has_mita', 'no')
            total_units = request.POST.get('total_units', '0')
            
            if first_name and phone_number and street_id:
                street = Street.objects.get(id=street_id)
                
                # Parse total_units - default to 0 if not provided or invalid
                try:
                    units_value = float(total_units) if has_mita == 'yes' and total_units else 0
                except (ValueError, TypeError):
                    units_value = 0
                
                Customer.objects.create(
                    first_name=first_name,
                    last_name=last_name or '',
                    phone_number=phone_number,
                    street=street,
                    total_units=units_value
                )
                messages.success(request, f'Customer "{first_name} {last_name}" added successfully.')
                return redirect('manager_customer_list')
            else:
                messages.error(request, 'Please fill in all required fields.')
    
    customers = Customer.objects.select_related('street').all()
    streets = Street.objects.all()
    
    # Calculate customer stats
    total_customers = customers.count()
    customers_with_units = customers.filter(total_units__gt=0).count()
    customers_without_units = customers.filter(total_units=0).count()
    
    return render(request, 'manager_customers.html', {
        'customers': customers,
        'streets': streets,
        'active_page': 'customers',
        'total_customers': total_customers,
        'customers_with_units': customers_with_units,
        'customers_without_units': customers_without_units,
    })


@login_required(login_url='login')
def manager_street_list(request):
    if request.user.role != 'manager':
        return redirect('login')
    
    if request.method == 'POST':
        region = request.POST.get('region')
        district = request.POST.get('district')
        street_name = request.POST.get('street_name')
        
        if region and district and street_name:
            Street.objects.get_or_create(
                region=region,
                district=district,
                street_name=street_name
            )
            messages.success(request, f'Street "{street_name}" added successfully.')
        else:
            messages.error(request, 'Please fill in all required fields.')
        return redirect('manager_street_list')
    
    streets = Street.objects.all()
    return render(request, 'manager_streets.html', {
        'streets': streets,
        'active_page': 'streets'
    })


@login_required(login_url='login')
def manager_sales_list(request):
    if request.user.role != 'manager':
        return redirect('login')
    
    # Get week filter from query param
    week_filter = request.GET.get('week', 'all')
    
    # Calculate available weeks
    today = timezone.now().date()
    current_week_start = today - timedelta(days=today.weekday())
    weeks = []
    for i in range(12):  # Last 12 weeks
        week_start = current_week_start - timedelta(weeks=i)
        week_end = week_start + timedelta(days=6)
        week_num = week_start.isocalendar()[1]
        weeks.append({
            'value': i,
            'label': f"Week {week_num} ({week_start.strftime('%b %d')} - {week_end.strftime('%b %d')})",
            'start': week_start,
            'end': week_end
        })
    
    # Filter orders based on week
    if week_filter != 'all':
        try:
            week_idx = int(week_filter)
            week_start = current_week_start - timedelta(weeks=week_idx)
            week_end = week_start + timedelta(days=6)
            orders = Order.objects.filter(
                created_at__date__gte=week_start,
                created_at__date__lte=week_end
            ).select_related('customer', 'street', 'created_by').order_by('-created_at')
        except ValueError:
            orders = Order.objects.select_related('customer', 'street', 'created_by').all().order_by('-created_at')
    else:
        orders = Order.objects.select_related('customer', 'street', 'created_by').all().order_by('-created_at')
    
    # Calculate totals
    totals = orders.aggregate(
        total_units=Sum('units'),
        total_revenue=Sum('total_amount')
    )
    
    return render(request, 'manager_sales.html', {
        'orders': orders,
        'total_units': totals['total_units'] or 0,
        'total_revenue': totals['total_revenue'] or 0,
        'active_page': 'sales',
        'weeks': weeks,
        'selected_week': week_filter
    })


@login_required(login_url='login')
def manager_expenditure_list(request):
    if request.user.role != 'manager':
        return redirect('login')
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'delete':
            expenditure_id = request.POST.get('expenditure_id')
            try:
                expenditure = Expenditure.objects.get(id=expenditure_id)
                expenditure.delete()
                messages.success(request, 'Expenditure deleted successfully.')
            except Expenditure.DoesNotExist:
                messages.error(request, 'Expenditure not found.')
            return redirect('manager_expenditure_list')
        
        # Add new expenditure
        date = request.POST.get('date')
        amount = request.POST.get('amount')
        purpose = request.POST.get('purpose')
        
        if date and amount and purpose:
            Expenditure.objects.create(
                date=date,
                amount=amount,
                purpose=purpose,
                created_by=request.user
            )
            messages.success(request, 'Expenditure added successfully.')
        else:
            messages.error(request, 'Please fill in all required fields.')
        return redirect('manager_expenditure_list')
    
    # Get week filter from query param
    week_filter = request.GET.get('week', 'all')
    
    # Calculate available weeks
    today = timezone.now().date()
    current_week_start = today - timedelta(days=today.weekday())
    weeks = []
    for i in range(12):  # Last 12 weeks
        week_start = current_week_start - timedelta(weeks=i)
        week_end = week_start + timedelta(days=6)
        week_num = week_start.isocalendar()[1]
        weeks.append({
            'value': i,
            'label': f"Week {week_num} ({week_start.strftime('%b %d')} - {week_end.strftime('%b %d')})",
            'start': week_start,
            'end': week_end
        })
    
    # Filter expenditures based on week
    if week_filter != 'all':
        try:
            week_idx = int(week_filter)
            week_start = current_week_start - timedelta(weeks=week_idx)
            week_end = week_start + timedelta(days=6)
            expenditures = Expenditure.objects.filter(
                date__gte=week_start,
                date__lte=week_end
            ).order_by('-created_at')
        except ValueError:
            expenditures = Expenditure.objects.all().order_by('-created_at')
    else:
        expenditures = Expenditure.objects.all().order_by('-created_at')
    
    # Calculate totals
    # Only count paid orders (exclude debt) for available amount
    total_revenue = Order.objects.exclude(payment_method='debt').aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
    total_debt = Order.objects.filter(payment_method='debt').aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
    # Add debt payments received to total revenue
    total_debt_payments = DebtPayment.objects.aggregate(total=Sum('amount'))['total'] or Decimal('0')
    total_income = total_revenue + total_debt_payments
    total_expenditure = expenditures.aggregate(total=Sum('amount'))['total'] or Decimal('0')
    amount_available = total_income - total_expenditure
    
    return render(request, 'manager_expenditure.html', {
        'expenditures': expenditures,
        'total_revenue': total_revenue,
        'total_debt_payments': total_debt_payments,
        'total_income': total_income,
        'total_expenditure': total_expenditure,
        'amount_available': amount_available,
        'active_page': 'expenditure',
        'weeks': weeks,
        'selected_week': week_filter
    })


@login_required(login_url='login')
def manager_weekly_report(request):
    if request.user.role != 'manager':
        return redirect('login')
    
    # Get week number from query param, default to current week
    week_offset = int(request.GET.get('week', 0))
    
    # Calculate current week's Monday
    today = timezone.now().date()
    current_week_start = today - timedelta(days=today.weekday())
    
    # Apply week offset (negative for past weeks)
    week_start = current_week_start - timedelta(weeks=week_offset)
    week_end = week_start + timedelta(days=6)
    
    # Calculate week number of the year
    week_number = week_start.isocalendar()[1]
    year = week_start.year
    
    # Get orders for this week
    weekly_orders = Order.objects.filter(
        created_at__date__gte=week_start,
        created_at__date__lte=week_end
    ).select_related('customer', 'street', 'created_by').order_by('-created_at')
    
    # Calculate order stats (all orders including debt)
    order_stats = weekly_orders.aggregate(
        total_units=Sum('units'),
        total_amount=Sum('total_amount')
    )
    total_units = order_stats['total_units'] or 0
    total_sales = order_stats['total_amount'] or 0
    total_orders = weekly_orders.count()
    
    # Calculate paid sales (exclude debt orders for profit calculation)
    paid_orders = weekly_orders.exclude(payment_method='debt')
    paid_sales_stats = paid_orders.aggregate(total_amount=Sum('total_amount'))
    paid_sales = paid_sales_stats['total_amount'] or Decimal('0')
    
    # Calculate debt sales for this week
    debt_orders = weekly_orders.filter(payment_method='debt')
    debt_sales_stats = debt_orders.aggregate(total_amount=Sum('total_amount'))
    total_debt_sales = debt_sales_stats['total_amount'] or Decimal('0')
    
    # Get debt payments received this week (from previous debts)
    debt_payments_this_week = DebtPayment.objects.filter(
        payment_date__gte=week_start,
        payment_date__lte=week_end
    ).aggregate(total=Sum('amount'))
    debt_payments_received = debt_payments_this_week['total'] or Decimal('0')
    
    # Get expenditures for this week
    weekly_expenditures = Expenditure.objects.filter(
        date__gte=week_start,
        date__lte=week_end
    ).aggregate(total=Sum('amount'))
    total_expenditure = weekly_expenditures['total'] or Decimal('0')
    
    # Calculate total income (paid sales + debt payments received)
    total_income = paid_sales + debt_payments_received
    
    # Calculate profit (paid sales + debt payments - expenditure)
    profit = total_income - total_expenditure
    
    return render(request, 'manager_weekly_report.html', {
        'active_page': 'weekly_report',
        'week_offset': week_offset,
        'week_number': week_number,
        'year': year,
        'week_start': week_start,
        'week_end': week_end,
        'total_units': total_units,
        'total_sales': total_sales,
        'paid_sales': paid_sales,
        'total_debt_sales': total_debt_sales,
        'debt_payments_received': debt_payments_received,
        'total_income': total_income,
        'total_orders': total_orders,
        'total_expenditure': total_expenditure,
        'profit': profit,
        'orders': weekly_orders,
    })


@login_required(login_url='login')
def manager_reports(request):
    if request.user.role != 'manager':
        return redirect('login')
    
    # Handle custom report entry form submission
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'add':
            mwaka = request.POST.get('mwaka')
            mwezi = request.POST.get('mwezi')
            units_sold = request.POST.get('units_sold', 0)
            sales = request.POST.get('sales', 0)
            expenditure = request.POST.get('expenditure', 0)
            
            if mwaka and mwezi:
                CustomReportEntry.objects.create(
                    mwaka=mwaka,
                    mwezi=mwezi,
                    units_sold=units_sold or 0,
                    sales=sales or 0,
                    expenditure=expenditure or 0,
                    created_by=request.user
                )
                messages.success(request, 'Custom report entry added successfully.')
            else:
                messages.error(request, 'Please fill in Mwaka and Mwezi fields.')
        
        elif action == 'delete':
            entry_id = request.POST.get('entry_id')
            if entry_id:
                CustomReportEntry.objects.filter(id=entry_id).delete()
                messages.success(request, 'Entry deleted successfully.')
        
        return redirect('manager_reports')
    
    # Get all custom report entries
    custom_entries = CustomReportEntry.objects.all()
    
    # Calculate custom entries totals
    custom_totals = custom_entries.aggregate(
        units_sold=Sum('units_sold'),
        sales=Sum('sales'),
        expenditure=Sum('expenditure')
    )
    custom_totals['units_sold'] = custom_totals['units_sold'] or 0
    custom_totals['sales'] = custom_totals['sales'] or 0
    custom_totals['expenditure'] = custom_totals['expenditure'] or 0
    custom_totals['profit'] = custom_totals['sales'] - custom_totals['expenditure']
    
    # Get selected year from query param, default to current year
    selected_year = int(request.GET.get('year', timezone.now().year))
    
    # Get available years from orders
    from django.db.models.functions import ExtractYear
    order_years = Order.objects.annotate(year=ExtractYear('created_at')).values_list('year', flat=True).distinct().order_by('-year')
    available_years = list(order_years) if order_years else [timezone.now().year]
    
    # Ensure current year is in the list
    current_year = timezone.now().year
    if current_year not in available_years:
        available_years.insert(0, current_year)
    
    # Generate weekly report data for the selected year
    from datetime import date
    import calendar
    
    weekly_data = []
    
    # Get first Monday of the year using ISO week calendar
    year_start = date(selected_year, 1, 1)
    # Find the Monday of the week containing Jan 1
    first_monday = year_start - timedelta(days=year_start.weekday())
    
    # If first_monday is in previous year, move to next week
    if first_monday.year < selected_year:
        first_monday = first_monday + timedelta(weeks=1)
    
    # Generate data for each week of the year
    week_start = first_monday
    
    while week_start.year == selected_year:
        week_end = week_start + timedelta(days=6)
        
        # Use ISO week number to match Funga Hesabu
        week_num = week_start.isocalendar()[1]
        
        # Stop if we've gone past the selected year
        if week_start.year > selected_year:
            break
        
        # Get orders for this week
        week_orders = Order.objects.filter(
            created_at__date__gte=week_start,
            created_at__date__lte=week_end
        )
        
        order_stats = week_orders.aggregate(
            total_units=Sum('units'),
            total_sales=Sum('total_amount')
        )
        
        # Get paid sales (exclude debt orders)
        paid_orders = week_orders.exclude(payment_method='debt')
        paid_sales_stats = paid_orders.aggregate(total_amount=Sum('total_amount'))
        paid_sales = paid_sales_stats['total_amount'] or Decimal('0')
        
        # Get debt sales for this week
        debt_orders = week_orders.filter(payment_method='debt')
        debt_sales_stats = debt_orders.aggregate(total_amount=Sum('total_amount'))
        debt_sales = debt_sales_stats['total_amount'] or Decimal('0')
        
        # Get debt payments received this week
        debt_payments_week = DebtPayment.objects.filter(
            payment_date__gte=week_start,
            payment_date__lte=week_end
        ).aggregate(total=Sum('amount'))
        debt_payments_received = debt_payments_week['total'] or Decimal('0')
        
        # Get expenditures for this week
        week_expenditure = Expenditure.objects.filter(
            date__gte=week_start,
            date__lte=week_end
        ).aggregate(total=Sum('amount'))
        
        units_sold = order_stats['total_units'] or 0
        sales = paid_sales + debt_payments_received  # Total income
        expenditure = week_expenditure['total'] or Decimal('0')
        profit = sales - expenditure
        
        # Only add weeks that have data or are in the past/current
        today = timezone.now().date()
        if units_sold > 0 or sales > 0 or expenditure > 0 or debt_sales > 0 or debt_payments_received > 0 or week_start <= today:
            weekly_data.append({
                'year': selected_year,
                'week_num': week_num,
                'week_start': week_start,
                'week_end': week_end,
                'units_sold': units_sold,
                'paid_sales': paid_sales,
                'debt_sales': debt_sales,
                'debt_payments_received': debt_payments_received,
                'sales': sales,
                'expenditure': expenditure,
                'profit': profit,
            })
        
        week_start = week_start + timedelta(weeks=1)
        
        # Safety check - max 53 weeks
        if len(weekly_data) > 53:
            break
    
    # Calculate yearly totals
    yearly_totals = {
        'units_sold': sum(w['units_sold'] for w in weekly_data),
        'paid_sales': sum(w['paid_sales'] for w in weekly_data),
        'debt_sales': sum(w['debt_sales'] for w in weekly_data),
        'debt_payments_received': sum(w['debt_payments_received'] for w in weekly_data),
        'sales': sum(w['sales'] for w in weekly_data),
        'expenditure': sum(w['expenditure'] for w in weekly_data),
        'profit': sum(w['profit'] for w in weekly_data),
    }
    
    # Calculate all-time system totals
    all_orders = Order.objects.all()
    all_order_stats = all_orders.aggregate(
        total_units=Sum('units'),
        total_sales=Sum('total_amount')
    )
    all_expenditure = Expenditure.objects.aggregate(total=Sum('amount'))
    
    system_totals = {
        'units_sold': all_order_stats['total_units'] or 0,
        'sales': all_order_stats['total_sales'] or 0,
        'expenditure': all_expenditure['total'] or 0,
    }
    system_totals['profit'] = system_totals['sales'] - system_totals['expenditure']
    
    # Calculate overall totals (custom entries + system data)
    overall_totals = {
        'units_sold': custom_totals['units_sold'] + system_totals['units_sold'],
        'sales': custom_totals['sales'] + system_totals['sales'],
        'expenditure': custom_totals['expenditure'] + system_totals['expenditure'],
        'profit': custom_totals['profit'] + system_totals['profit'],
    }
    
    return render(request, 'manager_reports.html', {
        'active_page': 'reports',
        'selected_year': selected_year,
        'available_years': available_years,
        'weekly_data': weekly_data,
        'yearly_totals': yearly_totals,
        'custom_entries': custom_entries,
        'custom_totals': custom_totals,
        'system_totals': system_totals,
        'overall_totals': overall_totals,
    })


@login_required(login_url='login')
def manager_water_production(request):
    if request.user.role != 'manager':
        return redirect('login')
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'delete':
            production_id = request.POST.get('production_id')
            try:
                production = WaterProduction.objects.get(id=production_id)
                production_date = production.date
                production.delete()
                messages.success(request, f'Rekodi ya tarehe {production_date} imefutwa.')
            except WaterProduction.DoesNotExist:
                messages.error(request, 'Rekodi haipatikani.')
            except Exception as e:
                messages.error(request, f'Error: {str(e)}')
            return redirect('manager_water_production')
        
        date = request.POST.get('date')
        units_produced = request.POST.get('units_produced')
        notes = request.POST.get('notes', '')
        
        if date and units_produced:
            try:
                production, created = WaterProduction.objects.update_or_create(
                    date=date,
                    defaults={
                        'units_produced': Decimal(units_produced),
                        'notes': notes,
                        'user': request.user
                    }
                )
                if created:
                    messages.success(request, f'Production record for {date} added successfully.')
                else:
                    messages.success(request, f'Production record for {date} updated successfully.')
            except Exception as e:
                messages.error(request, f'Error: {str(e)}')
        else:
            messages.error(request, 'Please fill in all required fields.')
        return redirect('manager_water_production')
    
    # Get filter parameters
    selected_month = request.GET.get('month', '')
    selected_year = request.GET.get('year', '')
    
    # Get all production records
    productions = WaterProduction.objects.all()
    
    # Apply month/year filter if provided
    if selected_month and selected_year:
        productions = productions.filter(
            date__month=int(selected_month),
            date__year=int(selected_year)
        )
    elif selected_year:
        productions = productions.filter(date__year=int(selected_year))
    
    # Calculate totals (filtered)
    total_produced = productions.aggregate(total=Sum('units_produced'))['total'] or Decimal('0')
    
    # Calculate sold units for the filtered period
    if selected_month and selected_year:
        total_sold = Order.objects.filter(
            created_at__month=int(selected_month),
            created_at__year=int(selected_year)
        ).aggregate(total=Sum('units'))['total'] or Decimal('0')
    elif selected_year:
        total_sold = Order.objects.filter(
            created_at__year=int(selected_year)
        ).aggregate(total=Sum('units'))['total'] or Decimal('0')
    else:
        total_sold = Order.objects.aggregate(total=Sum('units'))['total'] or Decimal('0')
    
    total_remaining = total_produced - total_sold
    
    # Get available years for filter dropdown
    available_years = WaterProduction.objects.dates('date', 'year', order='DESC')
    years = [d.year for d in available_years]
    if not years:
        years = [timezone.now().year]
    
    # Months list
    months = [
        (1, 'Januari'), (2, 'Februari'), (3, 'Machi'), (4, 'Aprili'),
        (5, 'Mei'), (6, 'Juni'), (7, 'Julai'), (8, 'Agosti'),
        (9, 'Septemba'), (10, 'Oktoba'), (11, 'Novemba'), (12, 'Desemba')
    ]
    
    return render(request, 'manager_water_production.html', {
        'active_page': 'water_production',
        'productions': productions,
        'total_produced': total_produced,
        'total_sold': total_sold,
        'total_remaining': total_remaining,
        'months': months,
        'years': years,
        'selected_month': selected_month,
        'selected_year': selected_year,
    })


@login_required(login_url='login')
def manager_customer_debts(request):
    if request.user.role != 'manager':
        return redirect('login')
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'delete':
            debt_id = request.POST.get('debt_id')
            try:
                debt = CustomerDebt.objects.get(id=debt_id)
                debt.delete()
                messages.success(request, 'Deni limefutwa.')
            except CustomerDebt.DoesNotExist:
                messages.error(request, 'Deni halipatikani.')
            return redirect('manager_customer_debts')
        
        if action == 'pay':
            debt_id = request.POST.get('debt_id')
            payment_amount = request.POST.get('payment_amount')
            try:
                debt = CustomerDebt.objects.get(id=debt_id)
                payment_decimal = Decimal(payment_amount)
                
                # Create payment record in history
                DebtPayment.objects.create(
                    debt=debt,
                    amount=payment_decimal,
                    received_by=request.user
                )
                
                # Update debt amount_paid
                debt.amount_paid += payment_decimal
                if debt.amount_paid >= debt.amount:
                    debt.is_paid = True
                debt.save()
                messages.success(request, 'Malipo yamerekodiwa.')
            except CustomerDebt.DoesNotExist:
                messages.error(request, 'Deni halipatikani.')
            return redirect('manager_customer_debts')
        
        # Add new debt
        customer_id = request.POST.get('customer')
        units = request.POST.get('units')
        amount = request.POST.get('amount')
        
        if customer_id and units and amount:
            try:
                customer = Customer.objects.get(id=customer_id)
                units_decimal = Decimal(units)
                amount_decimal = Decimal(amount)
                
                # Get active unit config for price
                active_config = UnitConfig.objects.filter(is_active=True).first()
                unit_price = active_config.price_per_unit if active_config else Decimal('5000')
                
                # Create the Order first (like a credit sale)
                order = Order.objects.create(
                    customer=customer,
                    street=customer.street,
                    units=units_decimal,
                    unit_price=unit_price,
                    total_amount=amount_decimal,
                    payment_method='debt',
                    created_by=request.user
                )
                
                # Create the debt record linked to the order
                CustomerDebt.objects.create(
                    customer=customer,
                    order=order,
                    units=units_decimal,
                    amount=amount_decimal,
                    created_by=request.user
                )
                
                # Update customer's total units (like an order)
                customer.total_units += units_decimal
                customer.save()
                
                messages.success(request, f'Deni limeongezwa. Order: {order.order_number}')
            except Customer.DoesNotExist:
                messages.error(request, 'Mteja hapatikani.')
        else:
            messages.error(request, 'Tafadhali jaza sehemu zote.')
        return redirect('manager_customer_debts')
    
    # Get all debts with order info
    debts = CustomerDebt.objects.select_related('customer', 'customer__street', 'order').prefetch_related('payments').all()
    
    # Calculate totals
    total_debts = debts.aggregate(total=Sum('amount'))['total'] or Decimal('0')
    total_units_debt = debts.aggregate(total=Sum('units'))['total'] or Decimal('0')
    total_paid = debts.aggregate(total=Sum('amount_paid'))['total'] or Decimal('0')
    total_remaining = total_debts - total_paid
    
    # Get customers for dropdown
    customers = Customer.objects.filter(is_active=True).order_by('first_name')
    
    # Get active unit config for price calculation
    active_config = UnitConfig.objects.filter(is_active=True).first()
    price_per_unit = active_config.price_per_unit if active_config else Decimal('5000')
    
    return render(request, 'manager_customer_debts.html', {
        'active_page': 'customer_debts',
        'debts': debts,
        'total_debts': total_debts,
        'total_units_debt': total_units_debt,
        'total_paid': total_paid,
        'total_remaining': total_remaining,
        'customers': customers,
        'price_per_unit': price_per_unit,
    })


@login_required(login_url='login')
def manager_get_debt_payments(request, debt_id):
    """API endpoint to get payment history for a debt"""
    if request.user.role != 'manager':
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    try:
        debt = CustomerDebt.objects.get(id=debt_id)
        payments = DebtPayment.objects.filter(debt=debt).order_by('-payment_date')
        
        payment_list = []
        for payment in payments:
            payment_list.append({
                'amount': float(payment.amount),
                'date': payment.payment_date.strftime('%d %b, %Y'),
                'received_by': payment.received_by.username if payment.received_by else '-',
                'notes': payment.notes or '-'
            })
        
        return JsonResponse({
            'payments': payment_list,
            'total_paid': float(debt.amount_paid),
            'total_debt': float(debt.amount),
            'remaining': float(debt.remaining_amount)
        })
    except CustomerDebt.DoesNotExist:
        return JsonResponse({'error': 'Debt not found'}, status=404)


@login_required(login_url='login')
def manager_reports_hub(request):
    if request.user.role != 'manager':
        return redirect('login')
    
    return render(request, 'manager_reports_hub.html', {
        'active_page': 'reports_hub',
    })


@login_required(login_url='login')
def manager_mitaa_report(request):
    if request.user.role != 'manager':
        return redirect('login')
    
    # Get all streets with their order statistics
    streets = Street.objects.all()
    
    street_data = []
    for street in streets:
        # Get total units sold and revenue for this street
        street_orders = Order.objects.filter(street=street)
        stats = street_orders.aggregate(
            total_units=Sum('units'),
            total_revenue=Sum('total_amount')
        )
        
        total_units = stats['total_units'] or Decimal('0')
        total_revenue = stats['total_revenue'] or Decimal('0')
        order_count = street_orders.count()
        customer_count = Customer.objects.filter(street=street).count()
        
        street_data.append({
            'name': street.street_name,
            'total_units': float(total_units),
            'total_revenue': float(total_revenue),
            'order_count': order_count,
            'customer_count': customer_count,
        })
    
    # Sort by total units (highest demand first)
    street_data.sort(key=lambda x: x['total_units'], reverse=True)
    
    # Prepare data for charts (JSON format)
    import json
    chart_labels = json.dumps([s['name'] for s in street_data])
    chart_units = json.dumps([s['total_units'] for s in street_data])
    chart_revenue = json.dumps([s['total_revenue'] for s in street_data])
    chart_orders = json.dumps([s['order_count'] for s in street_data])
    chart_customers = json.dumps([s['customer_count'] for s in street_data])
    
    # Calculate totals
    total_units_all = sum(s['total_units'] for s in street_data)
    total_revenue_all = sum(s['total_revenue'] for s in street_data)
    total_orders_all = sum(s['order_count'] for s in street_data)
    
    return render(request, 'manager_mitaa_report.html', {
        'active_page': 'reports_hub',
        'street_data': street_data,
        'chart_labels': chart_labels,
        'chart_units': chart_units,
        'chart_revenue': chart_revenue,
        'chart_orders': chart_orders,
        'chart_customers': chart_customers,
        'total_units_all': total_units_all,
        'total_revenue_all': total_revenue_all,
        'total_orders_all': total_orders_all,
    })


@login_required(login_url='login')
def manager_wateja_report(request):
    if request.user.role != 'manager':
        return redirect('login')
    
    # Get all streets with their customers
    streets = Street.objects.all().order_by('street_name')
    
    # Build customer data grouped by street
    street_customer_data = []
    all_customers = []
    
    for street in streets:
        customers = Customer.objects.filter(street=street).order_by('first_name')
        customer_list = []
        for idx, customer in enumerate(customers, 1):
            # Get customer order stats
            customer_orders = Order.objects.filter(customer=customer)
            stats = customer_orders.aggregate(
                total_units=Sum('units'),
                total_spent=Sum('total_amount')
            )
            
            customer_info = {
                'no': idx,
                'name': f"{customer.first_name} {customer.last_name}",
                'meter_no': float(customer.total_units or 0),
                'phone': customer.phone_number or '-',
                'total_units': float(stats['total_units'] or 0),
                'total_spent': float(stats['total_spent'] or 0),
            }
            customer_list.append(customer_info)
            all_customers.append({
                'street': street.street_name,
                **customer_info
            })
        
        if customer_list:
            street_customer_data.append({
                'street_name': street.street_name,
                'customers': customer_list,
                'customer_count': len(customer_list),
            })
    
    # Calculate totals for charts
    import json
    
    # Top customers by units purchased
    top_customers_by_units = sorted(all_customers, key=lambda x: x['total_units'], reverse=True)[:10]
    chart_top_names = json.dumps([c['name'][:15] + '...' if len(c['name']) > 15 else c['name'] for c in top_customers_by_units])
    chart_top_units = json.dumps([c['total_units'] for c in top_customers_by_units])
    
    # Customers per street
    chart_street_names = json.dumps([s['street_name'] for s in street_customer_data])
    chart_street_customers = json.dumps([s['customer_count'] for s in street_customer_data])
    
    # Calculate summary stats
    total_customers = len(all_customers)
    total_units_purchased = sum(c['total_units'] for c in all_customers)
    total_revenue = sum(c['total_spent'] for c in all_customers)
    
    return render(request, 'manager_wateja_report.html', {
        'active_page': 'reports_hub',
        'street_customer_data': street_customer_data,
        'all_customers': all_customers,
        'chart_top_names': chart_top_names,
        'chart_top_units': chart_top_units,
        'chart_street_names': chart_street_names,
        'chart_street_customers': chart_street_customers,
        'total_customers': total_customers,
        'total_units_purchased': total_units_purchased,
        'total_revenue': total_revenue,
    })


@login_required(login_url='login')
def manager_wateja_excel(request):
    if request.user.role != 'manager':
        return redirect('login')
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Wateja"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1E4A7A", end_color="1E4A7A", fill_type="solid")
    street_font = Font(bold=True, color="FFFFFF", size=11)
    street_fill = PatternFill(start_color="2D6CB7", end_color="2D6CB7", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    
    # Column headers
    headers = ['MITAA', 'NO.', 'MAJINA', 'MITA NO.', 'MAWASILIANO']
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
    
    # Set column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    
    # Get all streets with customers
    streets = Street.objects.all().order_by('street_name')
    
    row_num = 2
    for street in streets:
        customers = Customer.objects.filter(street=street).order_by('first_name')
        
        if not customers.exists():
            continue
        
        first_row_for_street = row_num
        
        for idx, customer in enumerate(customers, 1):
            # Street name (only on first row, will merge later)
            if idx == 1:
                ws.cell(row=row_num, column=1, value=street.street_name.upper())
            
            # Customer number
            cell = ws.cell(row=row_num, column=2, value=idx)
            cell.border = thin_border
            cell.alignment = center_align
            
            # Customer name
            cell = ws.cell(row=row_num, column=3, value=f"{customer.first_name} {customer.last_name}".upper())
            cell.border = thin_border
            cell.alignment = left_align
            
            # Meter number (total_units = current reading)
            cell = ws.cell(row=row_num, column=4, value=float(customer.total_units or 0))
            cell.border = thin_border
            cell.alignment = center_align
            
            # Phone number
            cell = ws.cell(row=row_num, column=5, value=customer.phone_number or '')
            cell.border = thin_border
            cell.alignment = center_align
            
            row_num += 1
        
        # Merge street name cells and apply style
        if row_num > first_row_for_street:
            last_row_for_street = row_num - 1
            if last_row_for_street > first_row_for_street:
                ws.merge_cells(start_row=first_row_for_street, start_column=1, 
                              end_row=last_row_for_street, end_column=1)
            
            # Style the street cell
            street_cell = ws.cell(row=first_row_for_street, column=1)
            street_cell.font = street_font
            street_cell.fill = street_fill
            street_cell.border = thin_border
            street_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Add empty row between streets
        row_num += 1
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="wateja_report.xlsx"'
    
    wb.save(response)
    return response


@login_required(login_url='login')
def manager_water_production_report(request):
    if request.user.role != 'manager':
        return redirect('login')
    
    import json
    from django.db.models.functions import TruncMonth, TruncWeek
    
    # Get all production records
    productions = WaterProduction.objects.all().order_by('date')
    
    # Calculate totals
    total_produced = WaterProduction.objects.aggregate(total=Sum('units_produced'))['total'] or Decimal('0')
    total_sold = Order.objects.aggregate(total=Sum('units'))['total'] or Decimal('0')
    total_remaining = total_produced - total_sold
    
    # Monthly production data
    monthly_data = WaterProduction.objects.annotate(
        month=TruncMonth('date')
    ).values('month').annotate(
        produced=Sum('units_produced')
    ).order_by('month')
    
    # Get monthly sales
    monthly_sales = Order.objects.annotate(
        month=TruncMonth('created_at')
    ).values('month').annotate(
        sold=Sum('units')
    ).order_by('month')
    
    # Build monthly chart data
    months_dict = {}
    for item in monthly_data:
        month_key = item['month'].strftime('%b %Y') if item['month'] else 'Unknown'
        months_dict[month_key] = {
            'produced': float(item['produced'] or 0),
            'sold': 0
        }
    
    for item in monthly_sales:
        month_key = item['month'].strftime('%b %Y') if item['month'] else 'Unknown'
        if month_key in months_dict:
            months_dict[month_key]['sold'] = float(item['sold'] or 0)
        else:
            months_dict[month_key] = {
                'produced': 0,
                'sold': float(item['sold'] or 0)
            }
    
    # Sort by date
    sorted_months = sorted(months_dict.keys(), key=lambda x: timezone.datetime.strptime(x, '%b %Y') if x != 'Unknown' else timezone.datetime.min)
    
    chart_months = json.dumps(sorted_months)
    chart_produced = json.dumps([months_dict[m]['produced'] for m in sorted_months])
    chart_sold = json.dumps([months_dict[m]['sold'] for m in sorted_months])
    
    # Daily production for last 30 days
    from datetime import timedelta
    thirty_days_ago = timezone.now().date() - timedelta(days=30)
    recent_productions = WaterProduction.objects.filter(date__gte=thirty_days_ago).order_by('date')
    
    daily_dates = json.dumps([p.date.strftime('%d %b') for p in recent_productions])
    daily_produced = json.dumps([float(p.units_produced) for p in recent_productions])
    daily_sold = json.dumps([float(p.units_sold) for p in recent_productions])
    
    # Production vs Sales comparison (pie chart)
    chart_comparison_labels = json.dumps(['Zilizouzwa', 'Zilizobaki'])
    chart_comparison_data = json.dumps([float(total_sold), float(total_remaining)])
    
    # Production records for table
    production_list = []
    for prod in productions:
        production_list.append({
            'date': prod.date,
            'produced': float(prod.units_produced),
            'sold': float(prod.units_sold),
            'remaining': float(prod.units_remaining),
            'notes': prod.notes or '-'
        })
    
    return render(request, 'manager_water_production_report.html', {
        'active_page': 'reports_hub',
        'total_produced': total_produced,
        'total_sold': total_sold,
        'total_remaining': total_remaining,
        'production_list': production_list,
        'chart_months': chart_months,
        'chart_produced': chart_produced,
        'chart_sold': chart_sold,
        'daily_dates': daily_dates,
        'daily_produced': daily_produced,
        'daily_sold': daily_sold,
        'chart_comparison_labels': chart_comparison_labels,
        'chart_comparison_data': chart_comparison_data,
    })


@login_required(login_url='login')
def manager_mapato_report(request):
    if request.user.role != 'manager':
        return redirect('login')
    
    import json
    from django.db.models.functions import TruncMonth, TruncWeek, ExtractYear, ExtractWeek
    
    # Get current year
    current_year = timezone.now().year
    
    # Get Custom Report Entries (Historical Data)
    custom_entries = CustomReportEntry.objects.all()
    custom_totals = custom_entries.aggregate(
        units_sold=Sum('units_sold'),
        sales=Sum('sales'),
        expenditure=Sum('expenditure')
    )
    historical_units = float(custom_totals['units_sold'] or 0)
    historical_mapato = float(custom_totals['sales'] or 0)
    historical_matumizi = float(custom_totals['expenditure'] or 0)
    historical_kipato = historical_mapato - historical_matumizi
    
    # Get yearly data
    yearly_data = []
    years = Order.objects.dates('created_at', 'year')
    
    for year_date in years:
        year = year_date.year
        year_orders = Order.objects.filter(created_at__year=year)
        year_expenditure = Expenditure.objects.filter(date__year=year)
        year_debts = CustomerDebt.objects.filter(date__year=year)
        year_debt_payments = DebtPayment.objects.filter(payment_date__year=year)
        
        total_units = year_orders.aggregate(total=Sum('units'))['total'] or Decimal('0')
        total_mapato = year_orders.aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
        total_matumizi = year_expenditure.aggregate(total=Sum('amount'))['total'] or Decimal('0')
        total_madeni_mapya = year_debts.aggregate(total=Sum('amount'))['total'] or Decimal('0')
        total_madeni_yaliyolipwa = year_debt_payments.aggregate(total=Sum('amount'))['total'] or Decimal('0')
        kipato = total_mapato - total_matumizi
        
        yearly_data.append({
            'year': year,
            'units': float(total_units),
            'mapato': float(total_mapato),
            'matumizi': float(total_matumizi),
            'madeni_mapya': float(total_madeni_mapya),
            'madeni_yaliyolipwa': float(total_madeni_yaliyolipwa),
            'kipato': float(kipato),
        })
    
    # Get weekly data for current year
    weekly_data = []
    for week_num in range(1, 53):
        # Calculate week start and end dates
        week_start = timezone.datetime.strptime(f'{current_year}-W{week_num:02d}-1', '%G-W%V-%u').date()
        week_end = week_start + timedelta(days=6)
        
        week_orders = Order.objects.filter(
            created_at__date__gte=week_start,
            created_at__date__lte=week_end
        )
        week_expenditure = Expenditure.objects.filter(
            date__gte=week_start,
            date__lte=week_end
        )
        week_debts = CustomerDebt.objects.filter(
            date__gte=week_start,
            date__lte=week_end
        )
        week_debt_payments = DebtPayment.objects.filter(
            payment_date__gte=week_start,
            payment_date__lte=week_end
        )
        
        total_units = week_orders.aggregate(total=Sum('units'))['total'] or Decimal('0')
        total_mapato = week_orders.aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
        total_matumizi = week_expenditure.aggregate(total=Sum('amount'))['total'] or Decimal('0')
        total_madeni_mapya = week_debts.aggregate(total=Sum('amount'))['total'] or Decimal('0')
        total_madeni_yaliyolipwa = week_debt_payments.aggregate(total=Sum('amount'))['total'] or Decimal('0')
        kipato = total_mapato - total_matumizi
        
        weekly_data.append({
            'week': week_num,
            'units': float(total_units),
            'mapato': float(total_mapato),
            'matumizi': float(total_matumizi),
            'madeni_mapya': float(total_madeni_mapya),
            'madeni_yaliyolipwa': float(total_madeni_yaliyolipwa),
            'kipato': float(kipato),
        })
    
    # Calculate totals (system data only)
    system_units = sum(y['units'] for y in yearly_data)
    system_mapato = sum(y['mapato'] for y in yearly_data)
    system_matumizi = sum(y['matumizi'] for y in yearly_data)
    total_madeni_mapya_all = sum(y['madeni_mapya'] for y in yearly_data)
    total_madeni_yaliyolipwa_all = sum(y['madeni_yaliyolipwa'] for y in yearly_data)
    system_kipato = system_mapato - system_matumizi
    
    # Overall totals (historical + system)
    total_units_all = historical_units + system_units
    total_mapato_all = historical_mapato + system_mapato
    total_matumizi_all = historical_matumizi + system_matumizi
    total_kipato_all = total_mapato_all - total_matumizi_all
    
    # Chart data - Monthly for current year
    monthly_orders = Order.objects.filter(created_at__year=current_year).annotate(
        month=TruncMonth('created_at')
    ).values('month').annotate(
        mapato=Sum('total_amount'),
        units=Sum('units')
    ).order_by('month')
    
    monthly_expenditure = Expenditure.objects.filter(date__year=current_year).annotate(
        month=TruncMonth('date')
    ).values('month').annotate(
        matumizi=Sum('amount')
    ).order_by('month')
    
    # Build monthly chart data
    months_dict = {}
    month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    
    for item in monthly_orders:
        if item['month']:
            month_key = item['month'].strftime('%b')
            months_dict[month_key] = {
                'mapato': float(item['mapato'] or 0),
                'matumizi': 0,
                'units': float(item['units'] or 0)
            }
    
    for item in monthly_expenditure:
        if item['month']:
            month_key = item['month'].strftime('%b')
            if month_key in months_dict:
                months_dict[month_key]['matumizi'] = float(item['matumizi'] or 0)
            else:
                months_dict[month_key] = {
                    'mapato': 0,
                    'matumizi': float(item['matumizi'] or 0),
                    'units': 0
                }
    
    # Ensure all months are present
    for month in month_names:
        if month not in months_dict:
            months_dict[month] = {'mapato': 0, 'matumizi': 0, 'units': 0}
    
    chart_months = json.dumps(month_names)
    chart_mapato = json.dumps([months_dict[m]['mapato'] for m in month_names])
    chart_matumizi = json.dumps([months_dict[m]['matumizi'] for m in month_names])
    chart_units = json.dumps([months_dict[m]['units'] for m in month_names])
    
    # Yearly comparison chart
    chart_years = json.dumps([str(y['year']) for y in yearly_data])
    chart_yearly_mapato = json.dumps([y['mapato'] for y in yearly_data])
    chart_yearly_matumizi = json.dumps([y['matumizi'] for y in yearly_data])
    chart_yearly_kipato = json.dumps([y['kipato'] for y in yearly_data])
    
    return render(request, 'manager_mapato_report.html', {
        'active_page': 'reports_hub',
        'current_year': current_year,
        'yearly_data': yearly_data,
        'weekly_data': weekly_data,
        'custom_entries': custom_entries,
        'historical_units': historical_units,
        'historical_mapato': historical_mapato,
        'historical_matumizi': historical_matumizi,
        'historical_kipato': historical_kipato,
        'total_units_all': total_units_all,
        'total_mapato_all': total_mapato_all,
        'total_matumizi_all': total_matumizi_all,
        'total_madeni_mapya_all': total_madeni_mapya_all,
        'total_madeni_yaliyolipwa_all': total_madeni_yaliyolipwa_all,
        'total_kipato_all': total_kipato_all,
        'chart_months': chart_months,
        'chart_mapato': chart_mapato,
        'chart_matumizi': chart_matumizi,
        'chart_units': chart_units,
        'chart_years': chart_years,
        'chart_yearly_mapato': chart_yearly_mapato,
        'chart_yearly_matumizi': chart_yearly_matumizi,
        'chart_yearly_kipato': chart_yearly_kipato,
    })


@login_required(login_url='login')
def manager_mapato_excel(request):
    if request.user.role != 'manager':
        return redirect('login')
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Uendeshaji Mradi"
    
    # Styles
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, color="000000", size=11)
    header_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    
    # Title
    ws.merge_cells('A1:H1')
    title_cell = ws.cell(row=1, column=1, value="UENDESHAJI MRADI WA MAJI")
    title_cell.font = title_font
    title_cell.alignment = center_align
    
    # Headers
    headers = ['MWAKA', 'MWEZI', 'UNITS UZWA', 'MAPATO', 'MATUMIZI', 'MADENI MAPYA', 'MADENI YALIYOLIPWA', 'KIPATO']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
    
    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 18
    
    row_num = 3
    grand_total_units = Decimal('0')
    grand_total_mapato = Decimal('0')
    grand_total_matumizi = Decimal('0')
    grand_total_madeni_mapya = Decimal('0')
    grand_total_madeni_yaliyolipwa = Decimal('0')
    
    # Add Custom Report Entries (Historical Data) first
    custom_entries = CustomReportEntry.objects.all()
    for entry in custom_entries:
        # Year
        cell = ws.cell(row=row_num, column=1, value=entry.mwaka)
        cell.border = thin_border
        cell.alignment = center_align
        
        # Period
        cell = ws.cell(row=row_num, column=2, value=entry.mwezi)
        cell.border = thin_border
        
        # Units
        cell = ws.cell(row=row_num, column=3, value=float(entry.units_sold))
        cell.border = thin_border
        cell.alignment = right_align
        cell.number_format = '#,##0.00'
        
        # Mapato
        cell = ws.cell(row=row_num, column=4, value=float(entry.sales))
        cell.border = thin_border
        cell.alignment = right_align
        cell.number_format = '#,##0.00'
        
        # Matumizi
        cell = ws.cell(row=row_num, column=5, value=float(entry.expenditure))
        cell.border = thin_border
        cell.alignment = right_align
        cell.number_format = '#,##0.00'
        
        # Madeni Mapya (empty for historical)
        cell = ws.cell(row=row_num, column=6, value='')
        cell.border = thin_border
        
        # Madeni Yaliyolipwa (empty for historical)
        cell = ws.cell(row=row_num, column=7, value='')
        cell.border = thin_border
        
        # Kipato
        cell = ws.cell(row=row_num, column=8, value=float(entry.profit))
        cell.border = thin_border
        cell.alignment = right_align
        cell.number_format = '#,##0.00'
        
        grand_total_units += entry.units_sold
        grand_total_mapato += entry.sales
        grand_total_matumizi += entry.expenditure
        
        row_num += 1
    
    # Get yearly data
    years = Order.objects.dates('created_at', 'year')
    current_year = timezone.now().year
    
    for year_date in years:
        year = year_date.year
        
        if year == current_year:
            # For current year, show weekly breakdown
            first_row_for_year = row_num
            
            for week_num in range(1, 53):
                week_start = timezone.datetime.strptime(f'{year}-W{week_num:02d}-1', '%G-W%V-%u').date()
                week_end = week_start + timedelta(days=6)
                
                week_orders = Order.objects.filter(
                    created_at__date__gte=week_start,
                    created_at__date__lte=week_end
                )
                week_expenditure = Expenditure.objects.filter(
                    date__gte=week_start,
                    date__lte=week_end
                )
                week_debts = CustomerDebt.objects.filter(
                    date__gte=week_start,
                    date__lte=week_end
                )
                week_debt_payments = DebtPayment.objects.filter(
                    payment_date__gte=week_start,
                    payment_date__lte=week_end
                )
                
                total_units = week_orders.aggregate(total=Sum('units'))['total'] or Decimal('0')
                total_mapato = week_orders.aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
                total_matumizi = week_expenditure.aggregate(total=Sum('amount'))['total'] or Decimal('0')
                total_madeni_mapya = week_debts.aggregate(total=Sum('amount'))['total'] or Decimal('0')
                total_madeni_yaliyolipwa = week_debt_payments.aggregate(total=Sum('amount'))['total'] or Decimal('0')
                kipato = total_mapato - total_matumizi
                
                # Only add row if there's data or it's within reasonable weeks
                if total_units > 0 or total_mapato > 0 or total_matumizi > 0 or week_num <= 52:
                    # Year column (only first row)
                    if row_num == first_row_for_year:
                        ws.cell(row=row_num, column=1, value=year)
                    
                    # Week
                    cell = ws.cell(row=row_num, column=2, value=f"Wiki {week_num}")
                    cell.border = thin_border
                    
                    # Units
                    cell = ws.cell(row=row_num, column=3, value=float(total_units) if total_units > 0 else '')
                    cell.border = thin_border
                    cell.alignment = right_align
                    cell.number_format = '#,##0.00'
                    
                    # Mapato
                    cell = ws.cell(row=row_num, column=4, value=float(total_mapato) if total_mapato > 0 else '')
                    cell.border = thin_border
                    cell.alignment = right_align
                    cell.number_format = '#,##0.00'
                    
                    # Matumizi
                    cell = ws.cell(row=row_num, column=5, value=float(total_matumizi) if total_matumizi > 0 else '')
                    cell.border = thin_border
                    cell.alignment = right_align
                    cell.number_format = '#,##0.00'
                    
                    # Madeni Mapya
                    cell = ws.cell(row=row_num, column=6, value=float(total_madeni_mapya) if total_madeni_mapya > 0 else '')
                    cell.border = thin_border
                    cell.alignment = right_align
                    cell.number_format = '#,##0.00'
                    
                    # Madeni Yaliyolipwa
                    cell = ws.cell(row=row_num, column=7, value=float(total_madeni_yaliyolipwa) if total_madeni_yaliyolipwa > 0 else '')
                    cell.border = thin_border
                    cell.alignment = right_align
                    cell.number_format = '#,##0.00'
                    
                    # Kipato
                    cell = ws.cell(row=row_num, column=8, value=float(kipato) if kipato != 0 else '-')
                    cell.border = thin_border
                    cell.alignment = right_align
                    if isinstance(cell.value, float):
                        cell.number_format = '#,##0.00'
                    
                    grand_total_units += total_units
                    grand_total_mapato += total_mapato
                    grand_total_matumizi += total_matumizi
                    grand_total_madeni_mapya += total_madeni_mapya
                    grand_total_madeni_yaliyolipwa += total_madeni_yaliyolipwa
                    
                    row_num += 1
            
            # Merge year cells
            if row_num > first_row_for_year:
                last_row_for_year = row_num - 1
                if last_row_for_year > first_row_for_year:
                    ws.merge_cells(start_row=first_row_for_year, start_column=1,
                                  end_row=last_row_for_year, end_column=1)
                year_cell = ws.cell(row=first_row_for_year, column=1)
                year_cell.border = thin_border
                year_cell.alignment = center_align
        else:
            # For past years, show yearly total
            year_orders = Order.objects.filter(created_at__year=year)
            year_expenditure = Expenditure.objects.filter(date__year=year)
            year_debts = CustomerDebt.objects.filter(date__year=year)
            year_debt_payments = DebtPayment.objects.filter(payment_date__year=year)
            
            total_units = year_orders.aggregate(total=Sum('units'))['total'] or Decimal('0')
            total_mapato = year_orders.aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
            total_matumizi = year_expenditure.aggregate(total=Sum('amount'))['total'] or Decimal('0')
            total_madeni_mapya = year_debts.aggregate(total=Sum('amount'))['total'] or Decimal('0')
            total_madeni_yaliyolipwa = year_debt_payments.aggregate(total=Sum('amount'))['total'] or Decimal('0')
            kipato = total_mapato - total_matumizi
            
            # Year
            cell = ws.cell(row=row_num, column=1, value=year)
            cell.border = thin_border
            cell.alignment = center_align
            
            # Period description
            cell = ws.cell(row=row_num, column=2, value=f"Jan. {year} - Dec. {year}")
            cell.border = thin_border
            
            # Units
            cell = ws.cell(row=row_num, column=3, value=float(total_units))
            cell.border = thin_border
            cell.alignment = right_align
            cell.number_format = '#,##0.00'
            
            # Mapato
            cell = ws.cell(row=row_num, column=4, value=float(total_mapato))
            cell.border = thin_border
            cell.alignment = right_align
            cell.number_format = '#,##0.00'
            
            # Matumizi
            cell = ws.cell(row=row_num, column=5, value=float(total_matumizi))
            cell.border = thin_border
            cell.alignment = right_align
            cell.number_format = '#,##0.00'
            
            # Madeni Mapya
            cell = ws.cell(row=row_num, column=6, value=float(total_madeni_mapya))
            cell.border = thin_border
            cell.alignment = right_align
            cell.number_format = '#,##0.00'
            
            # Madeni Yaliyolipwa
            cell = ws.cell(row=row_num, column=7, value=float(total_madeni_yaliyolipwa))
            cell.border = thin_border
            cell.alignment = right_align
            cell.number_format = '#,##0.00'
            
            # Kipato
            cell = ws.cell(row=row_num, column=8, value=float(kipato))
            cell.border = thin_border
            cell.alignment = right_align
            cell.number_format = '#,##0.00'
            
            grand_total_units += total_units
            grand_total_mapato += total_mapato
            grand_total_matumizi += total_matumizi
            grand_total_madeni_mapya += total_madeni_mapya
            grand_total_madeni_yaliyolipwa += total_madeni_yaliyolipwa
            
            row_num += 1
    
    # Grand totals row
    row_num += 1
    ws.cell(row=row_num, column=1, value='')
    ws.cell(row=row_num, column=2, value='')
    
    cell = ws.cell(row=row_num, column=3, value=float(grand_total_units))
    cell.font = Font(bold=True)
    cell.alignment = right_align
    cell.number_format = '#,##0.00'
    
    cell = ws.cell(row=row_num, column=4, value=float(grand_total_mapato))
    cell.font = Font(bold=True)
    cell.alignment = right_align
    cell.number_format = '#,##0.00'
    
    cell = ws.cell(row=row_num, column=5, value=float(grand_total_matumizi))
    cell.font = Font(bold=True)
    cell.alignment = right_align
    cell.number_format = '#,##0.00'
    
    cell = ws.cell(row=row_num, column=6, value=float(grand_total_madeni_mapya))
    cell.font = Font(bold=True)
    cell.alignment = right_align
    cell.number_format = '#,##0.00'
    
    cell = ws.cell(row=row_num, column=7, value=float(grand_total_madeni_yaliyolipwa))
    cell.font = Font(bold=True)
    cell.alignment = right_align
    cell.number_format = '#,##0.00'
    
    cell = ws.cell(row=row_num, column=8, value=float(grand_total_mapato - grand_total_matumizi))
    cell.font = Font(bold=True)
    cell.alignment = right_align
    cell.number_format = '#,##0.00'
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="uendeshaji_mradi_report.xlsx"'
    
    wb.save(response)
    return response


@login_required(login_url='login')
def manager_water_production_list(request):
    if request.user.role != 'manager':
        return redirect('login')
    
    if request.method == 'POST':
        date = request.POST.get('date')
        units_produced = request.POST.get('units_produced')
        notes = request.POST.get('notes', '')
        
        if date and units_produced:
            try:
                production, created = WaterProduction.objects.update_or_create(
                    date=date,
                    defaults={
                        'units_produced': Decimal(units_produced),
                        'notes': notes,
                        'user': request.user
                    }
                )
                if created:
                    messages.success(request, f'Production record for {date} added successfully.')
                else:
                    messages.success(request, f'Production record for {date} updated successfully.')
            except Exception as e:
                messages.error(request, f'Error: {str(e)}')
        else:
            messages.error(request, 'Please fill in all required fields.')
        return redirect('manager_water_production_list')
    
    # Get filter parameters
    selected_month = request.GET.get('month', '')
    selected_year = request.GET.get('year', '')
    
    # Get all production records
    productions = WaterProduction.objects.all()
    
    # Apply month/year filter if provided
    if selected_month and selected_year:
        productions = productions.filter(
            date__month=int(selected_month),
            date__year=int(selected_year)
        )
    elif selected_year:
        productions = productions.filter(date__year=int(selected_year))
    
    # Calculate totals (filtered)
    total_produced = productions.aggregate(total=Sum('units_produced'))['total'] or Decimal('0')
    
    # Calculate sold units for the filtered period
    if selected_month and selected_year:
        total_sold = Order.objects.filter(
            created_at__month=int(selected_month),
            created_at__year=int(selected_year)
        ).aggregate(total=Sum('units'))['total'] or Decimal('0')
    elif selected_year:
        total_sold = Order.objects.filter(
            created_at__year=int(selected_year)
        ).aggregate(total=Sum('units'))['total'] or Decimal('0')
    else:
        total_sold = Order.objects.aggregate(total=Sum('units'))['total'] or Decimal('0')
    
    total_remaining = total_produced - total_sold
    
    # Get available years for filter dropdown
    available_years = WaterProduction.objects.dates('date', 'year', order='DESC')
    years = [d.year for d in available_years]
    if not years:
        years = [timezone.now().year]
    
    # Months list
    months = [
        (1, 'Januari'), (2, 'Februari'), (3, 'Machi'), (4, 'Aprili'),
        (5, 'Mei'), (6, 'Juni'), (7, 'Julai'), (8, 'Agosti'),
        (9, 'Septemba'), (10, 'Oktoba'), (11, 'Novemba'), (12, 'Desemba')
    ]
    
    return render(request, 'manager_water_production.html', {
        'active_page': 'water_production',
        'productions': productions,
        'total_produced': total_produced,
        'total_sold': total_sold,
        'total_remaining': total_remaining,
        'months': months,
        'years': years,
        'selected_month': selected_month,
        'selected_year': selected_year,
    })


@login_required(login_url='login')
def download_customer_template(request):
    """Download Excel template for bulk customer import"""
    if request.user.role != 'manager':
        return redirect('login')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Wateja Template"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ['Jina la Kwanza', 'Jina la Mwisho', 'Namba ya Simu', 'Region', 'District', 'Mtaa (Street)', 'Mita No. (Units)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Set column widths
    column_widths = [20, 20, 18, 15, 15, 20, 18]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Add example row
    example_data = ['Juma', 'Hassan', '0712345678', 'Dar es Salaam', 'Kinondoni', 'Mwenge', '0']
    for col, value in enumerate(example_data, 1):
        cell = ws.cell(row=2, column=col, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
    
    # Add instructions row
    ws.cell(row=4, column=1, value="MAELEKEZO:")
    ws.cell(row=4, column=1).font = Font(bold=True, color="FF0000")
    ws.cell(row=5, column=1, value="1. Jina la Kwanza ni lazima (required)")
    ws.cell(row=6, column=1, value="2. Jina la Mwisho si lazima (optional)")
    ws.cell(row=7, column=1, value="3. Namba ya Simu ni lazima (required)")
    ws.cell(row=8, column=1, value="4. Region, District, na Mtaa - tumia orodha kwenye sheet 'Mitaa Iliyopo'")
    ws.cell(row=9, column=1, value="5. Mita No. weka 0 kama mteja hana mita")
    ws.cell(row=10, column=1, value="6. Futa mfano wa row 2 kabla ya kuupload")
    ws.cell(row=11, column=1, value="7. Angalia sheet 'Mitaa Iliyopo' kuona mitaa yote iliyopo")
    
    # Create second sheet with available streets
    ws_streets = wb.create_sheet(title="Mitaa Iliyopo")
    
    # Street sheet headers
    street_headers = ['S/N', 'Region', 'District', 'Mtaa (Street)']
    street_header_fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
    for col, header in enumerate(street_headers, 1):
        cell = ws_streets.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = street_header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Set street sheet column widths
    street_col_widths = [8, 20, 20, 25]
    for col, width in enumerate(street_col_widths, 1):
        ws_streets.column_dimensions[get_column_letter(col)].width = width
    
    # Add all available streets
    streets = Street.objects.all().order_by('region', 'district', 'street_name')
    for idx, street in enumerate(streets, 1):
        row_num = idx + 1
        ws_streets.cell(row=row_num, column=1, value=idx).border = thin_border
        ws_streets.cell(row=row_num, column=2, value=street.region).border = thin_border
        ws_streets.cell(row=row_num, column=3, value=street.district).border = thin_border
        ws_streets.cell(row=row_num, column=4, value=street.street_name).border = thin_border
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="wateja_template.xlsx"'
    wb.save(response)
    return response


@login_required(login_url='login')
def upload_customers_excel(request):
    """Upload Excel file to bulk import customers"""
    if request.user.role != 'manager':
        return redirect('login')
    
    if request.method != 'POST':
        return redirect('manager_customer_list')
    
    excel_file = request.FILES.get('excel_file')
    if not excel_file:
        messages.error(request, 'Tafadhali chagua faili ya Excel.')
        return redirect('manager_customer_list')
    
    if not excel_file.name.endswith(('.xlsx', '.xls')):
        messages.error(request, 'Tafadhali upload faili ya Excel (.xlsx au .xls).')
        return redirect('manager_customer_list')
    
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        success_count = 0
        error_count = 0
        errors = []
        
        # Skip header row, start from row 2
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Skip empty rows
            if not row or not any(row):
                continue
            
            first_name = str(row[0]).strip() if row[0] else ''
            last_name = str(row[1]).strip() if row[1] else ''
            phone_number = str(row[2]).strip() if row[2] else ''
            region = str(row[3]).strip() if row[3] else ''
            district = str(row[4]).strip() if row[4] else ''
            street_name = str(row[5]).strip() if row[5] else ''
            
            try:
                total_units = float(row[6]) if row[6] else 0
            except (ValueError, TypeError):
                total_units = 0
            
            # Validate required fields
            if not first_name:
                errors.append(f"Row {row_num}: Jina la Kwanza halijajazwa")
                error_count += 1
                continue
            
            if not phone_number:
                errors.append(f"Row {row_num}: Namba ya Simu haijajazwa")
                error_count += 1
                continue
            
            if not region or not district or not street_name:
                errors.append(f"Row {row_num}: Region, District, au Mtaa haijajazwa")
                error_count += 1
                continue
            
            # Get or create street
            street, _ = Street.objects.get_or_create(
                region=region,
                district=district,
                street_name=street_name
            )
            
            # Create customer
            Customer.objects.create(
                first_name=first_name,
                last_name=last_name,
                phone_number=phone_number,
                street=street,
                total_units=total_units
            )
            success_count += 1
        
        if success_count > 0:
            messages.success(request, f'Wateja {success_count} wameongezwa kwa mafanikio!')
        
        if error_count > 0:
            error_msg = f'Wateja {error_count} hawakuongezwa. '
            if errors[:5]:
                error_msg += 'Makosa: ' + '; '.join(errors[:5])
                if len(errors) > 5:
                    error_msg += f' ... na mengine {len(errors) - 5}'
            messages.warning(request, error_msg)
        
    except Exception as e:
        messages.error(request, f'Hitilafu wakati wa kusoma faili: {str(e)}')
    
    return redirect('manager_customer_list')
